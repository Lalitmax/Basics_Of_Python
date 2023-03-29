import openpyxl
import speech_recognition as sr
import pyttsx3
import time


day = int(input("Enter day 1,2,3... upto 30 : "))
eng = pyttsx3.init()


def speak(text):

    eng.say(text)
    eng.runAndWait()


def comaaaaand(name):

    r = sr.Recognizer()

    with sr.Microphone() as mic:
        r.adjust_for_ambient_noise(mic)
        r.pause_threshold = 0.8
        print(name)
        eng.say(name)
        eng.runAndWait()

        print("Listening...")
        audio = r.listen(mic)

    try:

        text = r.recognize_google(audio)
        text = text.lower()
    except Exception as ex:
        print(ex)
        speak("say again")
        return "None"
    return text


rm = openpyxl.load_workbook("Attendancer.xlsx")

sh1 = rm["Sheet1"]
sh2 = rm["Sheet2"]

length = sh1.max_row
# for write


# for read

li_name = []
for i in range(2, length+1):
    name = sh1.cell(i, 1).value

    rc = comaaaaand(name)

    if rc in ["present sir", "yes sir"]:
        sh2.cell(row=i, column=1+day, value="Present")
        rm.save("Attendancer.xlsx")

    else:
        sh2.cell(row=i, column=1+day, value="Not Present")
        rm.save("Attendancer.xlsx")


# print(li_name)
